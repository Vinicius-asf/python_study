############################################################
## Trabalho de Termodinâmica Computacional                ##
## Professor Marcio Ziviani                               ##
## Alunos                                                 ##
## Vinícius Antônio Soares Ferreira                       ##
## Marcelo Rocha de Castilho Sampaio                      ##
## Universidade Federal de Minas Gerais - 04 de Dezembro  ##
## 2018/2                                                 ##
############################################################

import numpy as np
from openpyxl import Workbook

###########################################################
## Declaração das matrizes, vetores e valores da simulação

A = {} # dicionario de coeficientes
B = {} # dicionario de termos independentes
T = {} # dicionario de temperaturas

deltaT = 0.0839 # tempo entre os cálculos

A[0] = np.array([ # coeficientes das equações de cada malha
  [-11.78, 4.68, 0, 2.344, 0, 0, 0, 0, 0, 0, 0, 0],         #1
  [4.6875, -11.78, 4.6875, 0, 2.344, 0, 0, 0, 0, 0, 0, 0],  #2
  [0, 9.375, -11.718, 0, 0, 2.344, 0, 0, 0, 0, 0, 0],       #3
  [1.172, 0, 0, -11.718, 4.6872, 0, 1.172, 0, 0, 0, 0, 0],  #4
  [0, 1.56, 0, 6.25, -11.92, 3.125, 0, 0.78, 0, 0, 0, 0],   #5
  [0, 0, 0, 0, 0, -11.91, 2.344, 0, 0, 0, 4.6875, 0],       #6
  [0, 0, 0, 1.172, 0, 0, -11.718, 4.68, 0, 1.172, 0, 0],    #7
  [0, 0, 0, 0, 0.78, 0, 6.25, -11.92, 3.125, 0, 1.56, 0],   #8
  [0, 0, 0, 0, 0, 0, 0, 9.375, -11.916, 0, 0, 2.344],       #9
  [0, 0, 0, 0, 0, 0, 2.344, 0, 0, -11.718, 4.69, 0],        #10
  [0, 0, 0, 0, 0, 0, 0, 2.344, 0, 4.68, -11.718, 4.68],     #11
  [0, 0, 0, 0, 0, 0, 0, 0, 2.344, 0, 9.375, -11.718]        #12
])

ident = np.identity(12) # matriz identidade que será somada no coeficiente

B[0] = np.array([ # termos independentes
  140,
  0, 
  0.0148, 
  140.62, 
  3.93, 
  3.96, 
  140.62, 
  3.93, 
  3.95, 
  190, 
  49.36, 
  49.28
])

T[0] = np.array([ # temperaturas no inicio - p = 0
  15,
  15,
  15,
  15,
  15,
  15,
  15,
  15,
  15,
  15,
  15,
  15,
])

finished = False
p = 0

####################################################
## Função e cálculo das temperaturas

def temperaturas(A,B,T,p): # função de calculo das temperaturas seguintes - Método explicito
  global deltaT, ident
  return ((A[0]*deltaT)+ident)@(T[p])+(B[0]*deltaT)

while not finished: 
  T[p+1] = temperaturas(A,B,T,p)
  if all(T[p+1]-T[p] < 0.001): # condição de parada: 0,001, para todas as diferenças entre p+1 e p
    finished = True
  p+=1

##########################################
## Print final e criação da pasta Excel

print('p= %i, Time= %.5f'%(p,p*deltaT))
print(T[p], T[p].shape)

wb = Workbook()
ws = wb.active

ws.cell(1,1,'p')
ws.cell(2,1,'tempo')
for i in range(1,13):
  ws.cell(i+2,1,i)

for p,temp in T.items():
  ws.cell(1,p+2,p)
  ws.cell(2,p+2,p*deltaT)
  for i in range(0,len(temp)):
    ws.cell(i+3,p+2,temp[i])

wb.save('Temperaturas.xlsx')
